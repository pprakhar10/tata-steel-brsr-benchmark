"""
generate_comparison_excel.py — Cross-company comparison table as Excel workbook

Run: python parser/generate_comparison_excel.py
Output: parser/review/comparison_table.xlsx  (3 sheets: FY2023, FY2024, FY2025)
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from generate_comparison_table import (
    TAG_GROUPS, COMPANIES, YEARS, load, get_val_obj, fmt_value, cell_class
)

OUT = Path(__file__).parent / "review" / "comparison_table.xlsx"

FILL_HEADER   = PatternFill("solid", fgColor="2C4A6E")
FILL_CATEGORY = PatternFill("solid", fgColor="D9E2ED")
FILL_PATCHED  = PatternFill("solid", fgColor="FFF3B0")
FILL_NULL     = PatternFill("solid", fgColor="FDE8E8")

FONT_HEADER   = Font(bold=True, color="FFFFFF", size=11)
FONT_CATEGORY = Font(bold=True, size=11)


def write_sheet(ws, year: str, company_data_map: dict):
    headers = ["Metric", "Unit"] + [name for _, name in COMPANIES] + ["Comments"]
    ws.append(headers)

    # Style header row
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    current_cat = None
    for category, tag, display_name, unit_hint in TAG_GROUPS:
        # Insert category row when category changes
        if category != current_cat:
            current_cat = category
            ws.append([category] + [""] * (len(headers) - 1))
            for cell in ws[ws.max_row]:
                cell.fill = FILL_CATEGORY
                cell.font = FONT_CATEGORY

        val_objs = [get_val_obj(company_data_map[cid], year, tag) for cid, _ in COMPANIES]
        row = [display_name, unit_hint or "—"] + [fmt_value(v) for v in val_objs] + [""]
        ws.append(row)

        # Color value cells (columns C–F = indices 3–6)
        data_row = ws[ws.max_row]
        for i, v_obj in enumerate(val_objs):
            cell = data_row[2 + i]  # 0-indexed: col 2 = Tata, 3 = JSW, 4 = SAIL, 5 = Jindal
            cls = cell_class(v_obj)
            if "patched" in cls:
                cell.fill = FILL_PATCHED
            elif "null-val" in cls:
                cell.fill = FILL_NULL

    # Set column widths
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    for col in ["C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 20
    ws.column_dimensions["G"].width = 35


def main():
    print("Loading company data...")
    company_data_map = {cid: load(cid) for cid, _ in COMPANIES}

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for year in YEARS:
        ws = wb.create_sheet(title=year)
        write_sheet(ws, year, company_data_map)
        print(f"  Sheet {year} done")

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    print(f"\nSaved: {OUT}")


if __name__ == "__main__":
    main()
