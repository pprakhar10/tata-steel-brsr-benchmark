"""
generate_analysis_review_excel.py — Analysis review workbook

Output: parser/review/analysis_review.xlsx
Columns A–U:
  A  KPI (display name)
  B  Principle
  C  Unit
  D  Direction           (green=clear, amber=UNCLEAR)
  E–H  FY2022-23 values  (Tata / JSW / SAIL / Jindal)
  I–L  FY2023-24 values
  M–P  FY2024-25 values
  Q  Tata Rank           (green=ever rank 1, amber=ever rank 4)
  R  Tata Trend
  S  AI Analysis         (light blue, all 6 sections, locked)
  T  Flags               (light orange when any flag is set)
  U  Changes Required    (yellow, editable)

Value cell highlights:
  Amber      = PDF-patched (corrected from source document)
  Light red  = null / not reported
  Light grey = explicit zero
  Orange border (left) = unit warning present
  Tata FY2022-23 = consolidated (noted in column header)

Flag highlights (column T):
  EXTERNAL_SOURCE_USED  = AI analysis cited sources beyond BRSR — ⚠ flagged
  CONTEXT_INSUFFICIENT  = insufficient data in BRSR for reliable analysis
  DIRECTION_UNCLEAR     = metric direction (higher/lower better) is ambiguous
  TREND_AMBIGUITY       = multi-year trend interpretation is uncertain

Run: python parser/generate_analysis_review_excel.py
"""

import json
from pathlib import Path

from openpyxl import Workbook
import re

from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# reuse TAG_GROUPS, company loader and value formatter from comparison table
import sys
sys.path.insert(0, str(Path(__file__).parent))
from generate_comparison_table import (
    TAG_GROUPS, COMPANIES, YEARS, load as load_company_data,
    get_val_obj, cell_class,
)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

BASE         = Path(__file__).parent.parent
ANALYSIS_DIR = BASE / "dashboard" / "src" / "data" / "analysis"
OUT_PATH     = BASE / "parser" / "review" / "analysis_review.xlsx"

# Year keys in analysis JSON match "FY2023" / "FY2024" / "FY2025"
# PEER_RANK also uses those keys
YEAR_KEYS = ["FY2023", "FY2024", "FY2025"]
YEAR_LABELS = {
    "FY2023": "FY2022-23",
    "FY2024": "FY2023-24",
    "FY2025": "FY2024-25",
}

PLACEHOLDER = "[ Analysis not yet generated ]"

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------

C_HEADER        = "2C4A6E"   # dark blue — column headers
C_CATEGORY      = "D9E2ED"   # light steel — category separator rows
C_PATCHED       = "FFF3B0"   # amber — PDF-corrected value
C_NULL          = "FDE8E8"   # light red — null / not reported
C_ZERO          = "F0F0F0"   # light grey — explicit zero
C_DIR_OK        = "C8E6C9"   # light green — direction is clear
C_DIR_UNCLEAR   = "FFE0B2"   # amber — direction unclear
C_RANK_BEST     = "C8E6C9"   # light green — Tata rank 1 in some year
C_RANK_WORST    = "FFE0B2"   # amber — Tata rank 4 in some year
C_AI_ANALYSIS   = "E8F4FD"   # light blue — AI analysis text
C_FLAGS         = "FFF0E0"   # light orange — flags set
C_FLAG_EXTERNAL = "FFD0A0"   # stronger orange — external source cited
C_CHANGES       = "FFFFF0"   # light yellow — editable changes column
C_CONSOLIDATED  = "FFF8DC"   # cornsilk — Tata FY2022-23 consolidated
C_UNIT_WARN     = "FF8C00"   # dark orange — used for left border only

FILL = {k: PatternFill("solid", fgColor=v) for k, v in {
    "header":       C_HEADER,
    "category":     C_CATEGORY,
    "patched":      C_PATCHED,
    "null":         C_NULL,
    "zero":         C_ZERO,
    "dir_ok":       C_DIR_OK,
    "dir_unclear":  C_DIR_UNCLEAR,
    "rank_best":    C_RANK_BEST,
    "rank_worst":   C_RANK_WORST,
    "ai_analysis":  C_AI_ANALYSIS,
    "flags":        C_FLAGS,
    "flag_ext":     C_FLAG_EXTERNAL,
    "changes":      C_CHANGES,
    "consolidated": C_CONSOLIDATED,
}.items()}

FONT_HEADER   = Font(bold=True, color="FFFFFF", size=10)
FONT_CATEGORY = Font(bold=True, size=10)
FONT_BOLD     = Font(bold=True, size=10)
FONT_NORMAL   = Font(size=10)
FONT_SMALL    = Font(size=9)

ALIGN_WRAP   = Alignment(wrap_text=True, vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")
ALIGN_TOP    = Alignment(vertical="top")

BORDER_UNIT_WARN = Border(
    left=Side(style="medium", color=C_UNIT_WARN)
)

# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------

def load_analysis(tag: str) -> dict | None:
    path = ANALYSIS_DIR / "indicators" / f"{tag}.json"
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# Formatters
# ---------------------------------------------------------------------------

def fmt_val_excel(v_obj) -> str:
    """Format value for Excel cell — no emoji, just the number."""
    if v_obj is None:
        return "—"
    val = v_obj.get("value")
    if val is None:
        return "—"
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if isinstance(val, (int, float)):
        av = abs(val)
        if av == 0:
            return "0"
        elif av >= 1_000_000:
            return f"{val:,.0f}"
        elif av >= 1_000:
            return f"{val:,.1f}"
        elif av >= 1:
            return f"{val:.4f}".rstrip("0").rstrip(".")
        else:
            return f"{val:.6g}"
    return str(val)


COMPANY_DISPLAY = [
    ("tata",   "Tata"),
    ("jsw",    "JSW"),
    ("sail",   "SAIL"),
    ("jindal", "Jindal"),
]


def fmt_all_ranks(analysis: dict) -> str:
    """All 4 companies x 3 years.
    FY2022-23: Tata=2, JSW=3, SAIL=1, Jindal=4
    FY2023-24: ...
    FY2024-25: ...
    """
    peer_rank = analysis.get("peerRank", {})
    lines = []
    for yr, label in zip(YEAR_KEYS, ["FY2022-23", "FY2023-24", "FY2024-25"]):
        yr_ranks = peer_rank.get(yr, {})
        parts = ", ".join(
            f"{name}={yr_ranks.get(key, '—')}"
            for key, name in COMPANY_DISPLAY
        )
        lines.append(f"{label}: {parts}")
    return "\n".join(lines)


def fmt_all_trends(analysis: dict) -> str:
    """All 4 companies, both periods.
    Tata:   FY23-24=Worsened, FY24-25=Stable
    JSW:    ...
    """
    td = analysis.get("trendDirection", {})
    lines = []
    for key, name in COMPANY_DISPLAY:
        co_td = td.get(key, {})
        fy2324 = co_td.get("FY2023-24", "—")
        fy2425 = co_td.get("FY2024-25", "—")
        lines.append(f"{name:<7} FY23-24={fy2324}, FY24-25={fy2425}")
    return "\n".join(lines)


def fmt_ai_analysis(analysis: dict) -> str:
    """Concatenate all 6 sections with plain-text headers."""
    sections = analysis.get("sections", {})
    parts = []
    for key, label in [
        ("about",                 "ABOUT"),
        ("performanceAndDrivers", "PERFORMANCE & KEY DRIVERS"),
        ("tataPositioning",       "TATA POSITIONING"),
        ("targets",               "TARGETS"),
        ("comparabilityNotes",    "COMPARABILITY NOTES"),
    ]:
        text = sections.get(key)
        if text:
            parts.append(f"{label}\n{text}")
    return "\n\n".join(parts)


def fmt_flags(analysis: dict) -> str:
    """
    Build the flags cell text.
    Prominently marks EXTERNAL_SOURCE_USED (info beyond BRSR).
    Returns empty string if no flags set.
    """
    flags = analysis.get("flags", {})
    lines = []

    if flags.get("externalSourceUsed"):
        lines.append("⚠ EXTERNAL SOURCE: Analysis cites information beyond BRSR disclosure")
    if flags.get("directionUnclear"):
        lines.append("⚠ DIRECTION UNCLEAR: Metric direction (higher/lower = better) is ambiguous")
    if flags.get("contextInsufficient"):
        lines.append("⚠ CONTEXT INSUFFICIENT: BRSR does not provide enough data for reliable analysis")
    if flags.get("trendAmbiguity"):
        lines.append("⚠ TREND AMBIGUITY: Multi-year trend interpretation is uncertain")

    note = flags.get("flagNote")
    if note and lines:
        lines.append(f"\nNote: {note}")

    return "\n".join(lines)


def fmt_year_cell(company_data_map: dict, tag: str, year: str) -> str:
    """
    Build compact multi-company value text for one year column.
    Tata   - 51,000 (C)
    JSW    - 38,500
    SAIL   - —
    Jindal - 850 (P)
    """
    lines = []
    for cid, cname, short in [
        ("tata-steel",       "Tata Steel",       "Tata  "),
        ("jsw-steel",        "JSW Steel",        "JSW   "),
        ("sail",             "SAIL",             "SAIL  "),
        ("jindal-stainless", "Jindal Stainless", "Jindal"),
    ]:
        v_obj = get_val_obj(company_data_map[cid], year, tag)
        val_str = fmt_val_excel(v_obj)
        markers = []
        if cid == "tata-steel" and year == "FY2023":
            note = (v_obj or {}).get("patchNote") or (v_obj or {}).get("unitWarning") or ""
            if re.search(r"\bindia\b|\bstandalone\b", note, re.IGNORECASE):
                markers.append("India")   # value is India/standalone basis
            else:
                markers.append("C")       # value is consolidated
        if v_obj and v_obj.get("patchSource") in ("PDF", "Excel"):
            markers.append("P")
        suffix = f" ({','.join(markers)})" if markers else ""
        lines.append(f"{short} - {val_str}{suffix}")
    return "\n".join(lines)


def tata_rank_color(analysis: dict) -> str | None:
    """Return color key for column Q: 'rank_best', 'rank_worst', or None."""
    peer_rank = analysis.get("peerRank", {})
    ranks = [peer_rank.get(yr, {}).get("tata") for yr in YEAR_KEYS]
    ranks = [r for r in ranks if r is not None]
    if not ranks:
        return None
    if 1 in ranks:
        return "rank_best"
    if 4 in ranks:
        return "rank_worst"
    return None


# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

# (header_text, width)
COLUMNS = [
    ("KPI",                              38),   # A
    ("Principle",                        14),   # B
    ("Unit",                             14),   # C
    ("Direction",                        18),   # D
    ("FY2022-23\n(C)=consolidated",      38),   # E
    ("FY2023-24",                        38),   # F
    ("FY2024-25",                        38),   # G
    ("All Ranks\n(1=best, 4=worst)",     32),   # H
    ("All Trends",                       44),   # I
    ("AI Analysis",                      80),   # J
    ("Flags\n(incl. sources beyond BRSR)", 34), # K
    ("Changes Required",                 38),   # L
]

N_COLS = len(COLUMNS)   # 12

# Column indices (1-based for openpyxl)
COL = {
    "kpi": 1, "principle": 2, "unit": 3, "direction": 4,
    "val_fy2023": 5,    # E
    "val_fy2024": 6,    # F
    "val_fy2025": 7,    # G
    "tata_rank":  8,    # H
    "tata_trend": 9,    # I
    "ai_analysis": 10,  # J
    "flags": 11,        # K
    "changes": 12,      # L
}

# Map year key → value column
VAL_COL = {"FY2023": COL["val_fy2023"], "FY2024": COL["val_fy2024"], "FY2025": COL["val_fy2025"]}

COMPANY_IDS = [cid for cid, _ in COMPANIES]


# ---------------------------------------------------------------------------
# Build sheet
# ---------------------------------------------------------------------------

def write_sheet(ws, company_data_map: dict):
    # ── Row 1: column headers ────────────────────────────────────────────────
    for col_idx, (header_text, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.fill = FILL["header"]
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER

    # ── Row 2: note on consolidated basis ───────────────────────────────────
    note_cell = ws.cell(row=2, column=COL["val_fy2023"],
                        value="(C) = Tata Steel FY2022-23 is Consolidated (includes international operations).  "
                              "(India) = Value confirmed as India/standalone basis (noted in manual review).  "
                              "(P) = Value patched/corrected from source PDF.  "
                              "All other company-years = Standalone (India only).")
    note_cell.font = Font(italic=True, size=9, color="885500")
    note_cell.fill = PatternFill("solid", fgColor=C_CONSOLIDATED)
    ws.merge_cells(start_row=2, start_column=COL["val_fy2023"],
                   end_row=2, end_column=N_COLS)

    # ── Data rows ────────────────────────────────────────────────────────────
    current_cat = None
    data_row = 3  # first data row (after header + note)

    for category, tag, display_name, unit_hint in TAG_GROUPS:

        # Category separator row
        if category != current_cat:
            current_cat = category
            cat_cell = ws.cell(row=data_row, column=1, value=category.upper())
            for col_idx in range(1, N_COLS + 1):
                c = ws.cell(row=data_row, column=col_idx)
                c.fill = FILL["category"]
                c.font = FONT_CATEGORY
            ws.merge_cells(start_row=data_row, start_column=1,
                           end_row=data_row, end_column=N_COLS)
            ws.row_dimensions[data_row].height = 18
            data_row += 1

        analysis = load_analysis(tag)

        # ── Columns A–C: KPI / Principle / Unit ─────────────────────────────
        ws.cell(row=data_row, column=1, value=display_name).font = FONT_NORMAL
        principle = (analysis or {}).get("principle", "")
        ws.cell(row=data_row, column=2, value=principle).font = FONT_SMALL
        ws.cell(row=data_row, column=3, value=unit_hint or "—").font = FONT_SMALL

        for col_idx in range(1, 4):
            ws.cell(row=data_row, column=col_idx).alignment = ALIGN_TOP

        # ── Column D: Direction ──────────────────────────────────────────────
        direction = (analysis or {}).get("direction", "")
        dir_cell = ws.cell(row=data_row, column=COL["direction"], value=direction)
        dir_cell.font = FONT_SMALL
        dir_cell.alignment = ALIGN_TOP
        if analysis:
            if direction.startswith("DIRECTION_UNCLEAR"):
                dir_cell.fill = FILL["dir_unclear"]
            elif direction in ("lower_is_better", "higher_is_better"):
                dir_cell.fill = FILL["dir_ok"]

        # ── Columns E–G: Values (one column per year, all 4 companies) ─────────
        for yr in YEAR_KEYS:
            col_idx = VAL_COL[yr]
            text = fmt_year_cell(company_data_map, tag, yr)
            cell = ws.cell(row=data_row, column=col_idx, value=text)
            cell.font = FONT_SMALL
            cell.alignment = ALIGN_WRAP
            # Tint FY2022-23 column to remind reviewer of consolidated basis
            if yr == "FY2023":
                cell.fill = PatternFill("solid", fgColor="FFFDF0")

        # ── Column Q: Tata Rank ──────────────────────────────────────────────
        q_cell = ws.cell(row=data_row, column=COL["tata_rank"])
        if analysis:
            q_cell.value = fmt_all_ranks(analysis)
            color_key = tata_rank_color(analysis)   # colour based on Tata's rank
            if color_key:
                q_cell.fill = FILL[color_key]
        else:
            q_cell.value = PLACEHOLDER
        q_cell.font = FONT_SMALL
        q_cell.alignment = ALIGN_WRAP

        # ── Column R: All Trends ─────────────────────────────────────────────
        r_cell = ws.cell(row=data_row, column=COL["tata_trend"])
        if analysis:
            r_cell.value = fmt_all_trends(analysis)
        else:
            r_cell.value = PLACEHOLDER
        r_cell.font = FONT_SMALL
        r_cell.alignment = ALIGN_WRAP

        # ── Column S: AI Analysis ────────────────────────────────────────────
        s_cell = ws.cell(row=data_row, column=COL["ai_analysis"])
        if analysis:
            s_cell.value = fmt_ai_analysis(analysis)
            s_cell.fill = FILL["ai_analysis"]
        else:
            s_cell.value = PLACEHOLDER
        s_cell.font = FONT_SMALL
        s_cell.alignment = ALIGN_WRAP

        # ── Column T: Flags ──────────────────────────────────────────────────
        flags_text = fmt_flags(analysis) if analysis else ""
        t_cell = ws.cell(row=data_row, column=COL["flags"], value=flags_text or "")
        t_cell.font = FONT_SMALL
        t_cell.alignment = ALIGN_WRAP
        if flags_text:
            flags_obj = (analysis or {}).get("flags", {})
            if flags_obj.get("externalSourceUsed"):
                t_cell.fill = FILL["flag_ext"]   # stronger orange = external source
            else:
                t_cell.fill = FILL["flags"]       # lighter orange = other flags

        # ── Column U: Changes Required ───────────────────────────────────────
        u_cell = ws.cell(row=data_row, column=COL["changes"], value="")
        u_cell.fill = FILL["changes"]
        u_cell.alignment = ALIGN_WRAP

        ws.row_dimensions[data_row].height = 90
        data_row += 1

    return data_row


# ---------------------------------------------------------------------------
# Legend sheet
# ---------------------------------------------------------------------------

def build_legend_sheet(wb):
    ws = wb.create_sheet(title="Legend")

    entries = [
        ("VALUE COLUMNS (E, F, G) — Text markers", None, None),
        ("(C) marker in cell",   C_CONSOLIDATED, "Tata Steel FY2022-23 is Consolidated (includes Indian + international operations). Not directly comparable to standalone peers."),
        ("(India) marker in cell", None,         "Value confirmed as India/standalone basis from manual review notes. Directly comparable to peers despite being in a consolidated filing year."),
        ("(P) marker in cell",   C_PATCHED,      "Value was corrected/patched from source PDF for comparability. Company annual report may show a different number."),
        ("— in cell",            C_NULL,         "Value not reported / data not available."),
        ("Light yellow column",  None,           "FY2022-23 column has a light tint as a reminder of the consolidated basis for Tata."),
        ("", None, None),
        ("CELL COLOURS — Analysis columns", None, None),
        ("Green (Direction col D)",   C_DIR_OK,    "Direction is clear (lower_is_better or higher_is_better)."),
        ("Amber (Direction col D)",   C_DIR_UNCLEAR,"Direction is ambiguous — AI flagged DIRECTION_UNCLEAR."),
        ("Green (Tata Rank col Q)",   C_RANK_BEST,  "Tata Steel ranked 1st (best) in at least one year."),
        ("Amber (Tata Rank col Q)",   C_RANK_WORST, "Tata Steel ranked 4th (worst) in at least one year."),
        ("Light blue (AI Analysis col S)", C_AI_ANALYSIS, "AI-generated analysis from Custom GPT. Locked — do not edit."),
        ("Strong orange (Flags col T)", C_FLAG_EXTERNAL, "⚠ EXTERNAL SOURCE: AI analysis cited information beyond BRSR disclosure. Verify the external source if needed."),
        ("Light orange (Flags col T)",  C_FLAGS,    "Other flags set: DIRECTION_UNCLEAR, CONTEXT_INSUFFICIENT, or TREND_AMBIGUITY."),
        ("Light yellow (Changes col U)", C_CHANGES, "Editable — enter your correction notes here."),
        ("", None, None),
        ("FLAGS — Column T", None, None),
        ("⚠ EXTERNAL SOURCE",        None, "AI analysis used information from outside the BRSR PDFs (e.g. company website, news, sustainability reports). The analysis is still valid but the external claim should be verified if being cited formally."),
        ("⚠ DIRECTION UNCLEAR",      None, "It is ambiguous whether a higher or lower value is better for this metric. The AI has flagged a reason."),
        ("⚠ CONTEXT INSUFFICIENT",   None, "The BRSR PDFs did not provide enough context for reliable analysis of this metric."),
        ("⚠ TREND AMBIGUITY",        None, "Multi-year trend comparison is uncertain (e.g. reporting methodology changed, one year is null, or boundary changed)."),
    ]

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 80

    ws.cell(row=1, column=1, value="Legend — analysis_review.xlsx").font = Font(bold=True, size=13)

    for i, (label, color, desc) in enumerate(entries, start=3):
        if not label:
            continue
        if color is None and desc is None:
            # section header
            cell = ws.cell(row=i, column=1, value=label)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill("solid", fgColor=C_CATEGORY)
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
            continue

        a = ws.cell(row=i, column=1, value=label)
        a.font = Font(size=10)
        a.alignment = Alignment(vertical="top")

        if color:
            b = ws.cell(row=i, column=2)
            b.fill = PatternFill("solid", fgColor=color)
            b.value = ""

        c = ws.cell(row=i, column=3, value=desc)
        c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 36


# ---------------------------------------------------------------------------
# Apply column widths, freeze panes, sheet protection
# ---------------------------------------------------------------------------

def apply_layout(ws):
    # Column widths
    for col_idx, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Header row height
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 22

    # Freeze: row 1 (header) + columns A–C (freeze at D3)
    ws.freeze_panes = "D3"

    # No sheet protection — allows free column/row resizing during review.
    # Column U yellow background is sufficient visual guidance for editable column.


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading company data...")
    company_data_map = {cid: load_company_data(cid) for cid, _ in COMPANIES}

    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis Review"
    ws.sheet_view.showGridLines = True

    print("Building analysis review sheet...")
    last_row = write_sheet(ws, company_data_map)

    print("Applying layout...")
    apply_layout(ws)

    print("Building legend sheet...")
    build_legend_sheet(wb)

    OUT_PATH.parent.mkdir(exist_ok=True)
    wb.save(OUT_PATH)

    # Summary
    ind_dir = ANALYSIS_DIR / "indicators"
    total_tags = len(TAG_GROUPS)
    analysis_count = sum(1 for _, tag, _, _ in TAG_GROUPS
                         if (ind_dir / f"{tag}.json").exists())
    flags_count = 0
    ext_source_count = 0
    for _, tag, _, _ in TAG_GROUPS:
        a = load_analysis(tag)
        if a:
            f = a.get("flags", {})
            if any(f.get(k) for k in ("directionUnclear", "contextInsufficient", "externalSourceUsed", "trendAmbiguity")):
                flags_count += 1
            if f.get("externalSourceUsed"):
                ext_source_count += 1

    print(f"\nWrote: {OUT_PATH}")
    print(f"  Rows: {total_tags} metrics + category separators")
    print(f"  Analysis populated: {analysis_count}/{total_tags}")
    print(f"  Indicators with any flag: {flags_count}")
    print(f"  External source cited (beyond BRSR): {ext_source_count}")
    print(f"\nReview workflow:")
    print(f"  1. Open analysis_review.xlsx")
    print(f"  2. Strong orange (col T) = external source cited — verify if needed")
    print(f"  3. Light orange (col T) = other flags — read the flag note")
    print(f"  4. Amber value cells = PDF-corrected; cornsilk = Tata consolidated")
    print(f"  5. Add corrections in col U (Changes Required) — only editable column")
    print(f"  6. Run apply_analysis_review.py to record corrections")


if __name__ == "__main__":
    main()
